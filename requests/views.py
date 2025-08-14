import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Subquery, OuterRef
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from .models import Request, RequestStatusHistory, RequestStatus
from .forms import RequestForm, ERPUpdateForm
from openpyxl.utils import get_column_letter

from accounts.models import UserDepartmentRight, Right, Department
from accounts.utils import has_department_right
from common.notify import notify_status_change

@login_required
def request_list(request):
    departments = Department.objects.all()
    statuses = RequestStatus.objects.all()
    user_depts = UserDepartmentRight.objects.filter(user=request.user).values_list('department', flat=True).distinct()

    requests_qs = Request.objects.filter(department__in=user_depts)
    has_requester_right = UserDepartmentRight.objects.filter(
        user=request.user,
        right__name='Requester'
    ).exists()
    # 🔍 Фильтры из GET
    dept_id = request.GET.get('department')
    month = request.GET.get('month')
    year = request.GET.get('year')
    status_code = request.GET.get('status')

    if dept_id:
        requests_qs = requests_qs.filter(department_id=dept_id)
    if month:
        requests_qs = requests_qs.filter(month=month)
    if year:
        requests_qs = requests_qs.filter(year=year)

    if status_code:
        last_status = RequestStatusHistory.objects.filter(
            request=OuterRef('pk')
        ).order_by('-changed_at').values('status__code')[:1]

        requests_qs = requests_qs.annotate(
            last_status_code=Subquery(last_status)
        ).filter(last_status_code=status_code)

    return render(request, 'requests/request_list.html', {
        'requests': requests_qs.order_by('-created_at'),
        'departments': departments,
        'statuses': statuses,
        'has_requester_right': has_requester_right,
        'filters': {
            'department': dept_id,
            'month': month,
            'year': year,
            'status': status_code,
        }
    })
@login_required
def request_create_or_edit(request, pk=None):
    instance = get_object_or_404(Request, pk=pk) if pk else None

    # if pk and instance.requester != request.user:
    #     return redirect('requests:request_list')
    if pk:
        # Только автор или апрувер в отделе
        print(instance.requester, request.user, instance.department)
        if not (
                instance.requester == request.user or
                has_department_right(request.user, instance.department, 'Approver')
        ):
            return redirect('requests:request_detail', pk=pk)
        last_status = instance.current_status()
        if last_status and last_status.status.code in ('approved', 'done', 'cancelled'):
            return redirect('requests:request_detail', pk=pk)
    else:
        has_right = UserDepartmentRight.objects.filter(
            user=request.user,
            right__name__in=['Requester', 'Approver']
        ).exists()
        if not has_right:
            return redirect('requests:request_list')

    if request.method == 'POST':
        form = RequestForm(request.POST, instance=instance)
        if form.is_valid():
            req = form.save(commit=False)
            req.requester = request.user

            if not has_department_right(request.user, req.department, 'Requester') and not has_department_right(request.user, req.department, 'Approver'):
                form.add_error(None, 'У вас нет прав на создание заявки в этом отделе.')
            else:
                req.save()

                # Только при создании (без pk ранее)
                if not pk:
                    draft_status = RequestStatus.objects.get(code='draft')
                    RequestStatusHistory.objects.create(
                        request=req,
                        status=draft_status,
                        changed_by=request.user
                    )

                return redirect('requests:request_list')
    else:
        form = RequestForm(instance=instance)



    return render(request, 'requests/request_form.html', {'form': form, 'edit': pk is not None})


@login_required
def request_detail(request, pk):
    req = get_object_or_404(Request, pk=pk)
    history = req.status_history.all().order_by('-changed_at')

    # Отладка — показать права текущего пользователя в отделе заявки
    user_rights = UserDepartmentRight.objects.filter(
        user=request.user,
        department=req.department
    )
    user_right_names = list(user_rights.values_list('right__name', flat=True))
    last_status = req.current_status()

    return render(request, 'requests/request_detail.html', {
        'request_obj': req,
        'history': history,
        'user_rights': user_rights,
        'user_right_names': user_right_names,
        'last_status': last_status,
    })

@login_required
def change_request_status(request, pk, new_status):
    req = get_object_or_404(Request, pk=pk)

    if not has_department_right(request.user, req.department, 'Approver'):
        return redirect('requests:request_detail', pk=pk)

    status_obj = get_object_or_404(RequestStatus, code=new_status)

    RequestStatusHistory.objects.create(
        request=req,
        status=status_obj,
        changed_by=request.user
    )

    # TODO: бизнес-логика: присваивание approver, обновление других флагов
    if new_status == 'approved' and req.approver is None:
        req.approver = request.user
        req.save()

    return redirect('requests:request_detail', pk=pk)

@login_required
def export_requests_excel(request):
    # фильтрация как в request_list
    user_depts = UserDepartmentRight.objects.filter(user=request.user).values_list('department', flat=True).distinct()
    requests_qs = Request.objects.filter(department__in=user_depts).order_by('-created_at')

    dept_id = request.GET.get('department')
    month = request.GET.get('month')
    year = request.GET.get('year')
    status_code = request.GET.get('status')

    if dept_id:
        requests_qs = requests_qs.filter(department_id=dept_id)
    if month:
        requests_qs = requests_qs.filter(month=month)
    if year:
        requests_qs = requests_qs.filter(year=year)

    if status_code:
        last_status = RequestStatusHistory.objects.filter(
            request=OuterRef('pk')
        ).order_by('-changed_at').values('status__code')[:1]

        requests_qs = requests_qs.annotate(
            last_status_code=Subquery(last_status)
        ).filter(last_status_code=status_code)

    # создание Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    columns = ['ID', 'Название', 'Отдел', 'Месяц', 'Год', 'Кол-во', 'Цена', 'Сумма', 'Статус', 'Заявитель']
    ws.append(columns)

    for row_idx, req in enumerate(requests_qs, start=2):
        status = req.current_status().status.description if req.current_status() else '—'
        ws.append([
            req.id,
            req.title,
            req.department.description,
            req.month,
            req.year,
            req.quantity,
            float(req.price),
            float(req.total),
            status,
            req.requester.username
        ])

    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # отдаём как файл
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=requests_export.xlsx'
    wb.save(response)
    return response

@login_required
def update_po(request, pk):
    req = get_object_or_404(Request, pk=pk)

    # ⛔ проверка прав перед любой логикой
    if not has_department_right(request.user, req.department, 'PO_manager'):
        return redirect('requests:request_detail', pk=pk)

    if request.method == 'POST':
        form = ERPUpdateForm(request.POST, instance=req)
        if form.is_valid():
            req = form.save()

            if req.current_status().status.code == 'approved':
                new_status = RequestStatus.objects.get(code='in_progress')
                RequestStatusHistory.objects.create(
                    request=req,
                    status=new_status,
                    changed_by=request.user
                )
                notify_status_change(req)

            messages.success(request, 'PO успешно обновлён.')
            return redirect('requests:request_detail', pk=pk)
    else:
        form = ERPUpdateForm(instance=req)

    return render(request, 'requests/update_po.html', {'form': form, 'request_obj': req})
